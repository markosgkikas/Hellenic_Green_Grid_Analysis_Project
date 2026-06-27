#!/usr/bin/env python3
"""
parse_admie.py — Parse the raw ADMIE Excel files into tidy long tables.

Each raw file is a human-formatted Excel sheet (Greek labels, multiple header
rows, a junk metadata sheet). This script normalizes every category into ONE
tidy schema:

    category, date, period, resolution, timestamp, series, value

  - period     : index within the day (1..24 hourly, 1..48 half-hourly),
                 empty for the reservoir snapshot
  - resolution : '60min' (legacy SCADA/market files), '30min' (Target Model
                 ISP files), or 'daily' (reservoir)
  - timestamp  : interval START = date + (period-1) * resolution
  - series     : the row label (unit / zone / country / load type / reservoir)
  - value      : numeric value

Files come in two on-disk formats and several structural families:
  format    .xls  -> xlrd ;  .xlsx -> openpyxl  (picked by extension)
  matrix    -> UnitProduction, SystemRealizationSCADA, DayAheadRESForecast,
               ISP1DayAheadRESForecast (periods across cols, series down rows)
  realtime  -> RealTimeSCADARES, RealTimeSCADASystemLoad, RealTimeSCADAImportsExports
  reservoir -> ReservoirFillingRate

Usage:
  python parse_admie.py --sample           # one file per category, show output
  python parse_admie.py                     # parse everything -> data/processed/<cat>.csv
  python parse_admie.py --filetypes ISP1DayAheadRESForecast -v
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
from datetime import datetime, timedelta

import xlrd
import openpyxl

RAW_ROOT = "data/raw/admie"
OUT_ROOT = "data/processed"

# Structural family per category.
FAMILY = {
    "UnitProduction": "matrix",
    "SystemRealizationSCADA": "matrix",
    "DayAheadRESForecast": "matrix",
    "ISP1DayAheadRESForecast": "matrix",   # Target Model successor (half-hourly)
    "RealTimeSCADARES": "realtime",
    "RealTimeSCADASystemLoad": "realtime",
    "RealTimeSCADAImportsExports": "realtime",
    "ReservoirFillingRate": "reservoir",
}

# Default resolution for the realtime/legacy families (always hourly). The
# matrix family auto-detects resolution per file from its period count, because
# ISP files changed from 30-min (48/day) to 15-min (96/day) over time.
DEFAULT_RES = ("60min", 60)


def infer_resolution(max_period):
    """Map a day's highest period number to (label, minutes-per-period).
    Tolerant of DST days: hourly 24/25, half-hourly 46/48/50, quarter 92/96/100."""
    if max_period <= 26:
        return ("60min", 60)
    if max_period <= 56:
        return ("30min", 30)
    return ("15min", 15)

ALL_CATEGORIES = list(FAMILY.keys())
OUT_COLUMNS = ["category", "date", "period", "resolution", "timestamp", "series", "value"]


# --------------------------------------------------------------------------- #
# small helpers
# --------------------------------------------------------------------------- #
def to_float(v):
    """Return v as float, or None if it isn't a usable number."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s in ("", "-", "N/A", "n/a", "#"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def as_period_int(v, max_period=100):
    """If v looks like a period label (1..max_period), return the int.

    Cap is 100 to cover 15-minute resolution (96/day, up to 100 on a DST
    fall-back day) as well as hourly (24/25) and half-hourly (48/50)."""
    f = to_float(v)
    if f is None:
        return None
    if abs(f - round(f)) < 1e-9 and 1 <= round(f) <= max_period:
        return int(round(f))
    return None


def period_header_map(row, min_len=12):
    """If row is a period header, return {column_index: period_number}.

    A header is a row whose period-like cells start at 1 and strictly increase
    left-to-right (>= min_len of them). Mapping by true period number tolerates
    stray gaps in the header (some ADMIE files drop an empty merged cell mid-row)
    and keeps data columns aligned to the right period even when periods are
    non-contiguous. Returns None if the row is not a header."""
    pairs = [(c, as_period_int(cell)) for c, cell in enumerate(row)]
    pairs = [(c, p) for c, p in pairs if p is not None]
    if len(pairs) < min_len:
        return None
    periods = [p for _, p in pairs]
    if periods[0] != 1:
        return None
    if any(b <= a for a, b in zip(periods, periods[1:])):  # strictly increasing
        return None
    return dict(pairs)


def clean_label(v):
    if not isinstance(v, str):
        return ""
    return re.sub(r"\s+", " ", v).strip()


def row_label(row, offset):
    """Join the non-empty text cells left of the first period column. Captures
    e.g. 'RES forecast / IPTO' (ISP) or the single series label (legacy)."""
    parts = [clean_label(row[c]) for c in range(min(offset, len(row)))]
    return " / ".join(p for p in parts if p)


def date_from_filename(path):
    """Extract a date (YYYY-MM-DD) from a raw ADMIE filename. Handles
    'YYYYMMDD_Category_NN.ext' and 'reportYYYYMMDD.xls'."""
    name = os.path.basename(path)
    m = re.search(r"(\d{8})", name)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def make_timestamp(d, period, minutes):
    base = datetime(d.year, d.month, d.day)
    return (base + timedelta(minutes=(period - 1) * minutes)).isoformat(sep=" ")


# --------------------------------------------------------------------------- #
# loading: pick engine by extension, choose the data sheet by content
# --------------------------------------------------------------------------- #
# Sheets that never hold data; their position vs. the data sheet drifts between
# years, so we pick the data sheet by content, not by index.
JUNK_SHEETS = {"xdo_metadata"}


def _load_xls(path):
    book = xlrd.open_workbook(path)
    sheets = [s for s in book.sheets() if s.name.strip().lower() not in JUNK_SHEETS]
    if not sheets:
        sheets = book.sheets()
    sh = max(sheets, key=lambda s: s.nrows * s.ncols)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def _load_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in wb.worksheets if s.title.strip().lower() not in JUNK_SHEETS]
    if not sheets:
        sheets = wb.worksheets
    ws = max(sheets, key=lambda s: (s.max_row or 0) * (s.max_column or 0))
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grid


def load_grid(path):
    return _load_xlsx(path) if path.lower().endswith(".xlsx") else _load_xls(path)


# --------------------------------------------------------------------------- #
# family parsers -> list of dict rows
# --------------------------------------------------------------------------- #
def _emit(rows, category, d, period, res_label, minutes, series, value):
    rows.append(
        {
            "category": category,
            "date": d.isoformat(),
            "period": period,
            "resolution": res_label,
            "timestamp": make_timestamp(d, period, minutes),
            "series": series,
            "value": value,
        }
    )


def parse_matrix(grid, category, d, res_label, minutes):
    """Periods across columns; series labels down rows. Multiple stacked blocks
    each introduced by a period-header row."""
    rows = []
    i, n = 0, len(grid)
    while i < n:
        hdr = period_header_map(grid[i])
        if not hdr:
            i += 1
            continue
        offset = min(hdr)
        cols = sorted(hdr.items())  # [(col, period), ...]
        # Resolution is detected from this block's period count: ISP forecasts
        # moved from 30-min (48/day) to 15-min (96/day) over time.
        res_label, minutes = infer_resolution(max(hdr.values()))
        i += 1
        while i < n:
            row = grid[i]
            if period_header_map(row):  # next block starts
                break
            label = row_label(row, offset)
            buf = []
            for col, period in cols:
                v = to_float(row[col]) if col < len(row) else None
                if v is None:
                    continue
                # 25th column on hourly files is the DST spillover slot;
                # 0.0 there is a placeholder on normal 24-hour days.
                if minutes == 60 and period == 25 and v == 0.0:
                    continue
                buf.append((period, v))
            if label and buf:
                for period, v in buf:
                    _emit(rows, category, d, period, res_label, minutes, label, v)
            i += 1
    return rows


def parse_realtime(grid, category, d, res_label, minutes):
    """Stacked blocks: a text label, then a 'Date' + periods header row, then a
    data row (date + hourly values)."""
    rows = []
    n = len(grid)
    last_label = ""
    for i, row in enumerate(grid):
        for cell in row:
            t = clean_label(cell)
            if t and t.lower() != "date" and to_float(cell) is None:
                last_label = t
                break

        date_col = next(
            (c for c, cell in enumerate(row) if clean_label(cell).lower() == "date"),
            None,
        )
        if date_col is None:
            continue
        sub = period_header_map(row[date_col + 1 :])
        if not sub:
            continue
        colmap = {date_col + 1 + c: p for c, p in sub.items()}
        if i + 1 >= n:
            continue
        drow = grid[i + 1]
        label = last_label or category
        for col, period in sorted(colmap.items()):
            v = to_float(drow[col]) if col < len(drow) else None
            if v is None:
                continue
            if minutes == 60 and period == 25 and v == 0.0:
                continue
            _emit(rows, category, d, period, res_label, minutes, label, v)
    return rows


def parse_reservoir(grid, category, d, res_label, minutes):
    """Snapshot: one filling-rate value per reservoir (no periods)."""
    rows = []
    header_row = ent_col = val_col = None
    for i, row in enumerate(grid):
        for c, cell in enumerate(row):
            if clean_label(cell).lower() == "entity":
                header_row, ent_col = i, c
                for c2 in range(c + 1, len(row)):
                    if clean_label(row[c2]):
                        val_col = c2
                        break
                break
        if header_row is not None:
            break
    if header_row is None:
        return rows
    if val_col is None:
        val_col = ent_col + 1
    for row in grid[header_row + 1 :]:
        entity = clean_label(row[ent_col]) if ent_col < len(row) else ""
        if not entity:
            continue
        val = to_float(row[val_col]) if val_col < len(row) else None
        if val is None:
            continue
        rows.append(
            {
                "category": category,
                "date": d.isoformat(),
                "period": "",
                "resolution": "daily",
                "timestamp": "",
                "series": entity,
                "value": val,
            }
        )
    return rows


PARSERS = {
    "matrix": parse_matrix,
    "realtime": parse_realtime,
    "reservoir": parse_reservoir,
}


def parse_file(path, category):
    d = date_from_filename(path)
    if d is None:
        raise ValueError(f"could not parse date from {os.path.basename(path)}")
    res_label, minutes = DEFAULT_RES  # matrix family overrides this per file
    grid = load_grid(path)
    return PARSERS[FAMILY[category]](grid, category, d, res_label, minutes)


# --------------------------------------------------------------------------- #
# file discovery
# --------------------------------------------------------------------------- #
def list_files(category):
    """All parseable Excel files for a category. UnitProduction lives as daily
    reportYYYYMMDD.xls files inside the extracted zip folders."""
    base = os.path.join(RAW_ROOT, category)
    if category == "UnitProduction":
        return sorted(glob.glob(os.path.join(base, "*", "report*.xls")))
    files = glob.glob(os.path.join(base, "*.xls")) + glob.glob(
        os.path.join(base, "*.xlsx")
    )
    return sorted(p for p in files if not p.endswith(".part"))


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #
def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Parse raw ADMIE Excel into tidy CSVs.")
    p.add_argument("--filetypes", nargs="+", default=ALL_CATEGORIES)
    p.add_argument(
        "--sample",
        action="store_true",
        help="Parse only ONE file per category and show the result (no full run).",
    )
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args(argv)


def write_csv(path, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # utf-8-sig writes a BOM so Excel on Windows auto-detects UTF-8 and renders
    # the Greek series labels correctly (plain utf-8 shows up as mojibake).
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=OUT_COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def run_sample(categories):
    for cat in categories:
        files = list_files(cat)
        if not files:
            print(f"[{cat}] no files found")
            continue
        path = files[-1]
        try:
            rows = parse_file(path, cat)
        except Exception as e:
            print(f"[{cat}] ERROR on {os.path.basename(path)}: {e}")
            continue
        out = os.path.join(OUT_ROOT, "_samples", f"{cat}_sample.csv")
        write_csv(out, rows)
        series = sorted({r["series"] for r in rows})
        res = rows[0]["resolution"] if rows else "?"
        print(f"\n[{cat}]  file={os.path.basename(path)}  resolution={res}")
        print(f"   tidy rows : {len(rows)}")
        print(f"   series ({len(series)}): {', '.join(series[:6])}{' ...' if len(series) > 6 else ''}")
        for r in rows[:3]:
            print(f"      {r['timestamp'] or r['date']}  p={r['period']}  {r['series']!r} = {r['value']}")
        print(f"   -> {out}")


def run_full(categories, verbose):
    grand = 0
    for cat in categories:
        files = list_files(cat)
        all_rows, ok, bad = [], 0, 0
        for path in files:
            try:
                all_rows.extend(parse_file(path, cat))
                ok += 1
            except Exception as e:
                bad += 1
                if verbose:
                    print(f"  [{cat}] FAILED {os.path.basename(path)}: {e}")
        out = os.path.join(OUT_ROOT, f"{cat}.csv")
        write_csv(out, all_rows)
        grand += len(all_rows)
        print(f"[{cat}] files ok={ok} bad={bad} -> {len(all_rows)} rows -> {out}")
    print(f"\nDONE. {grand} total tidy rows across {len(categories)} categories.")


def main(argv=None):
    args = parse_args(argv)
    for stream in (sys.stdout, sys.stderr):
        rc = getattr(stream, "reconfigure", None)
        if rc:
            try:
                rc(encoding="utf-8", errors="replace")
            except (ValueError, OSError):
                pass
    if args.sample:
        run_sample(args.filetypes)
    else:
        run_full(args.filetypes, args.verbose)
    return 0


if __name__ == "__main__":
    sys.exit(main())
